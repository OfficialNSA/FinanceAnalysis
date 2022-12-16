# Financial Analysis, creates a an excel report out of csv files for you to cross check financial decisions
# Copyright (C) 2022  Josua Gunzenhauser
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

# Contact the creator via chromsuport@gmail.com


# License and warranty hint in the running program
print("""Financial Analysis  Copyright (C) 2022  Josua Gunzenhauser
This program comes with ABSOLUTELY NO WARRANTY; for details see ./COPYING or <https://www.gnu.org/licenses/>
This is free software, and you are welcome to redistribute it
under certain conditions; see ./COPYING or <https://www.gnu.org/licenses/> for details.


""")

from fileinput import filename
import string
import pandas as pd
import locale
import json
import os

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart.series import DataPoint
from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference,
    LineChart
)
from openpyxl.chart.axis import DateAxis

def categorize_row(row):

    assigned = False

    if(row["IBAN Zahlungsbeteiligter"] in ignore_receivers):
        row["category"] = "ignore"
        assigned = True

    if(not assigned):
        for reason in reasons.keys():
            if(str(reason).lower() in str(row["Verwendungszweck"]).lower()):
                row["category"] = reasons[reason]
                assigned = True
                break

    if(not assigned):
        for receiver in receivers.keys():
            if(str(receiver).lower() in str(row["Name Zahlungsbeteiligter"]).lower()):
                row["category"] = receivers[receiver]
                assigned = True
                break

    if(not assigned):
        #Ask for entering new
        print("Eintrag muss kategorisiert werden:")
        pd.options.display.max_colwidth = 150
        print(row.drop(labels=["Bezeichnung Auftragskonto", "IBAN Auftragskonto", "BIC Auftragskonto", "Bankname Auftragskonto", "Buchungstag", "IBAN Zahlungsbeteiligter", "BIC (SWIFT-Code) Zahlungsbeteiligter", "Waehrung", "Saldo nach Buchung", "Bemerkung", "Kategorie", "Steuerrelevant", "Glaeubiger ID", "Mandatsreferenz"]))
        print("\nWelche Kategorie hat dieser Umsatz? Bitte den Index angeben oder neue Kategorie tippen")
        category_input = input(str(list(zip(range(len(categories)), categories))) + "\n")

        if(category_input == ""):
            category_input = "0"

        try:
            #Set the category from existing (index entered)
            row["category"] = categories[int(category_input)]
            category_input = categories[int(category_input)]
        except:
            #Set the category to the entered string and add it to categories
            row["category"] = category_input
            categories.append(category_input)

        type_input = input("Soll zukünftig der Verwendungszweck oder der Empfänger zur Zuordnung genutzt werden? [V]/[E]. [Enter]: nicht merken\n")

        if(type_input.lower().startswith("e") or type_input.lower().startswith("v")):

            term_input = input("Welcher Text soll zur Zuordnung genutzt werden? (Achte darauf, dass dieser nicht ungewollt oft in Wörtern vorkommt)\nNur[Enter]: Genau so wie Empfänger/Verwendungszweck\n")

            if(term_input == ""):
                if(type_input.lower().startswith("e")):
                    receivers[row["Name Zahlungsbeteiligter"]] = category_input
                elif(type_input.lower().startswith("v")):
                    reasons[row["Verwendungszweck"]] = category_input  
            else:
                if(type_input.lower().startswith("e")):
                    receivers[term_input] = category_input
                elif(type_input.lower().startswith("v")):
                    reasons[term_input] = category_input            
                
        print("\n")

    return row

def analyse_month(data):

    #Delete all unwanted information so that it doesn't appear after grouping
    data = data.drop(['Bezeichnung Auftragskonto', 'IBAN Auftragskonto', 'BIC Auftragskonto', 'Bankname Auftragskonto', 'Buchungstag', 'Valutadatum', 'Name Zahlungsbeteiligter', 'IBAN Zahlungsbeteiligter', 'BIC (SWIFT-Code) Zahlungsbeteiligter', 'Buchungstext', 'Verwendungszweck', 'Waehrung', 'Saldo nach Buchung', 'Bemerkung', 'Kategorie', 'Steuerrelevant', 'Glaeubiger ID', 'Mandatsreferenz'], axis=1)

    data = data.groupby("category").sum().sort_values("Betrag").drop("ignore")
    income = data[data["Betrag"] > 0]
    cost = data[data["Betrag"] < 0]

    return_value = {"income": income, "cost": cost}

    book = load_workbook("Analyse.xlsx")
    writer = pd.ExcelWriter("Analyse.xlsx", engine="openpyxl")
    writer.book = book

    income.to_excel(writer, sheet_name=filename + "_Analyse", startrow=0, startcol=0)
    cost.to_excel(writer, sheet_name=filename + "_Analyse", startrow=0, startcol=2)

    book = writer.book

    analysesheet = book[filename + "_Analyse"]

    pie_income = PieChart()
    labels_income = Reference(analysesheet, min_col=1, min_row=2, max_row=(1+income.shape[0]))
    data_income = Reference(analysesheet, min_col=2, min_row=1, max_row=(1+income.shape[0]))
    pie_income.add_data(data_income, titles_from_data=True)
    pie_income.set_categories(labels_income)
    pie_income.title = "Einnahmen"

    analysesheet.add_chart(pie_income, "F2")

    pie_cost = PieChart()
    labels_cost = Reference(analysesheet, min_col=3, min_row=2, max_row=(1+cost.shape[0]))
    data_cost = Reference(analysesheet, min_col=4, min_row=1, max_row=(1+cost.shape[0]))
    pie_cost.add_data(data_cost, titles_from_data=True)
    pie_cost.set_categories(labels_cost)
    pie_cost.title = "Ausgaben"

    analysesheet.add_chart(pie_cost, "F18")

    analysesheet.column_dimensions[get_column_letter(1)].auto_size = True
    analysesheet.column_dimensions[get_column_letter(2)].auto_size = True
    analysesheet.column_dimensions[get_column_letter(3)].auto_size = True
    analysesheet.column_dimensions[get_column_letter(4)].auto_size = True

    #Move the new analysis to the second page (first will be overall analysis)
    sheets=book._sheets

    #Bring the newest sheet to the front (pop from end, insert in front)
    sheet = sheets.pop(len(book._sheets) - 1)
    sheets.insert(1, sheet)

    book.save("Analyse.xlsx")

    writer.close()

    return return_value

def category_filter(month_data, category_data):
    
    #Delete all unwanted information so that it doesn't appear after grouping
    month_data = month_data.drop(['Bezeichnung Auftragskonto', 'IBAN Auftragskonto', 'BIC Auftragskonto', 'Bankname Auftragskonto', 'Valutadatum', 'BIC (SWIFT-Code) Zahlungsbeteiligter', 'Buchungstext', 'Waehrung', 'Saldo nach Buchung', 'Bemerkung', 'Kategorie', 'Steuerrelevant', 'Glaeubiger ID', 'Mandatsreferenz'], axis=1)

    unique_categories = list(month_data["category"].unique())
    unique_categories.remove("ignore")

    for category in unique_categories:
        try:
            category_data[category] = pd.concat([category_data[category], month_data[month_data["category"] == category]], ignore_index=True)
        except KeyError:
            category_data[category] = month_data[month_data["category"] == category]
        
    return category_data


#For European number notation
locale.setlocale(locale.LC_NUMERIC, "de")

# Create the Excel from scratch, makes the software easier to understand/maintain
try:
    os.remove("Analyse.xlsx")
except:
    pass

book = Workbook()
book.save("Analyse.xlsx")

# Create the Excel from scratch, makes the software easier to understand/maintain
try:
    os.remove("Kategorien.xlsx")
except:
    pass

book = Workbook()
book.save("Kategorien.xlsx")

# Create files and directories if they are not there already

if not os.path.isfile('categories.json'):
    categories_file = open('categories.json', 'w')
    json.dump({"categories": [],"receiver": {},"reason": {}}, categories_file)
    categories_file.close()

if not os.path.isfile('ignore.json'):
    categories_file = open('categories.json', 'w')
    json.dump({"receiverIBAN": [],"reason": []}, categories_file)
    categories_file.close()

if not os.path.isdir('Bank Exports'):
    os.mkdir("Bank Exports")

if not os.path.isdir('Categorized Exports'):
    os.mkdir("Categorized Exports")

#Load already assigned categories
categories_file = open('categories.json')
categories_content = json.load(categories_file)
categories_file.close()
categories = categories_content["categories"]
receivers = categories_content["receiver"]
reasons = categories_content["reason"]

#Load IBANs to ignore
ignore_file = open('ignore.json')
ignore_content = json.load(ignore_file)
ignore_file.close()
ignore_receivers = ignore_content["receiverIBAN"]
ignore_reasons = ignore_content["reason"]

#Go through all Exports and categorize them per prompt or per already assigned category if not already done previously
directory = os.fsencode("Bank Exports")

for file in os.listdir(directory):
    
    filename = os.fsdecode(file)

    if(not os.path.isfile(f"Categorized Exports/{filename}")):

        month_data = pd.read_csv(f"Bank Exports/{filename}", sep=";", decimal=",")

        month_data = month_data.apply(categorize_row, axis=1)

        month_data.to_csv(f"Categorized Exports/{filename}", index=None)
    
    else:
        print(f"{filename} already exists, no need to analyse")

#Write added categories back to json for next time
categories_content["categories"] = categories
categories_content["receiver"] = receivers
categories_content["reason"] = reasons
categories_file = open('categories.json', "w")
json.dump(categories_content, categories_file)
categories_file.close()

directory = os.fsencode("Categorized Exports")

complete_data = {}
category_data = {}

for file in os.listdir(directory):
    
    filename = os.fsdecode(file)

    month_data = pd.read_csv(f"Categorized Exports/{filename}")

    category_data = category_filter(month_data, category_data)

    complete_data[filename] = analyse_month(month_data)
    
# Here ok
# print(complete_data["2021.11.csv"]["income"])

sum = pd.DataFrame()
month_sums = []

#Abort if there is no data
if len(complete_data) == 0:
    print("\nEs konnten keine Daten gefunden werden, bitte Die Bankauszüge im csv Format in 'Bank Exports' ablegen")
    exit()

book = load_workbook("Kategorien.xlsx")
writer = pd.ExcelWriter("Kategorien.xlsx", engine="openpyxl")
writer.book = book

for category in category_data.keys():
    
    single_category_data = category_data[category]

    single_category_data = single_category_data.sort_values("Betrag")
    
    columns_titles = ["Buchungstag", "Name Zahlungsbeteiligter", "Betrag", "Verwendungszweck", "IBAN Zahlungsbeteiligter", "category"]
    single_category_data = single_category_data.reindex(columns=columns_titles)
    
    single_category_data.to_excel(writer, sheet_name=category, startrow=0, startcol=0, index=None)
    
    book = writer.book
    
    # Default Sheet delete
    try:
        del book['Sheet']
    except KeyError:
        pass
    
    sheet = book[category]

    for i in range(1, 7):
        sheet.column_dimensions[get_column_letter(i)].auto_size = True

book.save("Kategorien.xlsx")

writer.close()

for key in complete_data.keys():
    
    # Here ok
    # print(complete_data["2021.11.csv"]["income"])

    # Sum up cost and income for every month to compare in timeline
    month_sums.append(pd.Series(data={
        "income":complete_data[key]["income"].sum().get("Betrag"),
        "cost":complete_data[key]["cost"].sum().get("Betrag")},
        name=key.replace(".csv", "")))

    if(sum.empty):
        # Fill the DataFrame with some data (income chosen arbitrary)
        sum = complete_data[key]["income"]

        # Add the rest of the data per category. If the category isn't there yet, append it
        for category in complete_data[key]["cost"].index.values:
            try:
                sum.loc[category] = sum.loc[category] + complete_data[key]["cost"].loc[category]
            except KeyError:
                sum.loc[category] = complete_data[key]["cost"].loc[category]
                
    else:
        # Add the rest of the data per category. If the category isn't there yet, append it
        for category in complete_data[key]["income"].index.values:
            try:
                sum.loc[category] = sum.loc[category] + complete_data[key]["income"].loc[category]
            except KeyError:
                sum.loc[category] = complete_data[key]["income"].loc[category]

        # Add the rest of the data per category. If the category isn't there yet, append it
        for category in complete_data[key]["cost"].index.values:
            try:
                sum.loc[category] = sum.loc[category] + complete_data[key]["cost"].loc[category]
            except KeyError:
                sum.loc[category] = complete_data[key]["cost"].loc[category]

    # Here not ok
    # print(complete_data[key]["income"])

month_sums = pd.DataFrame(month_sums)

month_sums["cost"] = -month_sums["cost"]

sum = sum.groupby("category").sum().sort_values("Betrag")
sum_income = sum[sum["Betrag"] > 0]
sum_cost = sum[sum["Betrag"] < 0]

avg_sum = sum / len(complete_data.keys())

book = load_workbook("Analyse.xlsx")
writer = pd.ExcelWriter("Analyse.xlsx", engine="openpyxl")
writer.book = book

sum_income.to_excel(writer, sheet_name="Komplett_Analyse", startrow=0, startcol=0)
sum_cost.to_excel(writer, sheet_name="Komplett_Analyse", startrow=0, startcol=2)
month_sums.to_excel(writer, sheet_name="Komplett_Analyse", startrow=0, startcol=15)
avg_sum.to_excel(writer, sheet_name="Komplett_Analyse", startrow=max(sum_income.shape[0], sum_cost.shape[0])+4, startcol=1)

book = writer.book

analysesheet = book["Komplett_Analyse"]
analysesheet.cell(row=max(sum_income.shape[0], sum_cost.shape[0])+3, column=2).value = "Monatlich"

pie_income = PieChart()
labels_income = Reference(analysesheet, min_col=1, min_row=2, max_row=(1+sum_income.shape[0]))
data_income = Reference(analysesheet, min_col=2, min_row=1, max_row=(1+sum_income.shape[0]))
pie_income.add_data(data_income, titles_from_data=True)
pie_income.set_categories(labels_income)
pie_income.title = "Einnahmen"

analysesheet.add_chart(pie_income, "F2")

pie_cost = PieChart()
labels_cost = Reference(analysesheet, min_col=3, min_row=2, max_row=(1+sum_cost.shape[0]))
data_cost = Reference(analysesheet, min_col=4, min_row=1, max_row=(1+sum_cost.shape[0]))
pie_cost.add_data(data_cost, titles_from_data=True)
pie_cost.set_categories(labels_cost)
pie_cost.title = "Ausgaben"

analysesheet.add_chart(pie_cost, "F18")

timeline = LineChart()
timeline.title = "Verlauf"
timeline.style = 12
timeline.y_axis.title = ""
timeline.y_axis.crossAx = 500
timeline.x_axis = DateAxis(crossAx=100)
timeline.x_axis.number_format = 'yyyy.mm'
timeline.x_axis.majorTimeUnit = "months"
timeline.x_axis.title = "Datum"

labels_monthsum = Reference(analysesheet, min_col=16, min_row=2, max_row=(1+month_sums.shape[0]))
data_monthsum = Reference(analysesheet, min_col=17, max_col=18, min_row=1, max_row=(1+month_sums.shape[0]))
timeline.add_data(data_monthsum, titles_from_data=True)
timeline.set_categories(labels_monthsum)

analysesheet.add_chart(timeline, "P15")

analysesheet.column_dimensions[get_column_letter(1)].auto_size = True
analysesheet.column_dimensions[get_column_letter(2)].auto_size = True
analysesheet.column_dimensions[get_column_letter(3)].auto_size = True
analysesheet.column_dimensions[get_column_letter(4)].auto_size = True

#Move the new analysis to the second page (first will be overall analysis)
sheets=book._sheets

#Analysis Sheet
sheet = sheets.pop(len(book._sheets) - 1)
sheets.insert(1, sheet)

#Remove the empty default sheet from creation
book.remove(book.active)

book.save("Analyse.xlsx")

writer.close()
